"""XLSX export generator for dossier data.

Creates a multi-sheet Excel workbook with formatted financial data.
All openpyxl imports are deferred to avoid startup crashes if not installed.
"""

from io import BytesIO
from typing import Any


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, dict)):
        if isinstance(v, list):
            return ", ".join(str(x) for x in v[:10])
        entries = []
        for k, val in v.items():
            if isinstance(val, (dict, list)):
                entries.append(f"{k}: ...")
            else:
                entries.append(f"{k}: {val}")
        return "; ".join(entries[:8])
    return str(v)


def _flatten_dict(d: dict, prefix: str = "") -> list[tuple[str, Any]]:
    """Flatten a nested dict into (label, value) pairs."""
    rows: list[tuple[str, Any]] = []
    for k, v in d.items():
        if k.startswith("_"):
            continue
        label = f"{prefix}{k}".replace("_", " ").title()
        if isinstance(v, dict):
            if "value" in v or "current" in v or "score" in v:
                scalar = v.get("value", v.get("current", v.get("score")))
                rows.append((label, scalar))
            else:
                rows.extend(_flatten_dict(v, prefix=f"{k} > "))
        elif isinstance(v, list):
            rows.append((label, ", ".join(str(x) for x in v[:10])))
        else:
            rows.append((label, v))
    return rows


def generate_dossier_xlsx(dossier_data: dict) -> bytes:
    """Generate a multi-sheet XLSX workbook from dossier data.

    Returns the workbook as bytes suitable for StreamingResponse.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Style tokens ──
    HEADER_FILL = PatternFill(start_color="0C0E12", end_color="0C0E12", fill_type="solid")
    HEADER_FONT = Font(name="Arial", bold=True, color="3B82F6", size=10)
    DATA_FONT = Font(name="Consolas", color="D1D5DB", size=10)
    LABEL_FONT = Font(name="Arial", color="9CA3AF", size=10)
    TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    SUBTITLE_FONT = Font(name="Arial", color="6B7280", size=10)
    CELL_BORDER = Border(bottom=Side(style="thin", color="1F2937"))
    ALT_ROW_FILL = PatternFill(start_color="13161B", end_color="13161B", fill_type="solid")

    def write_header(ws, row: int, title: str, col_count: int = 2) -> int:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name="Arial", bold=True, color="3B82F6", size=11)
        cell.fill = HEADER_FILL
        for c in range(2, col_count + 1):
            ws.cell(row=row, column=c).fill = HEADER_FILL
        return row + 1

    def write_kv_rows(ws, start_row: int, data: list[tuple[str, Any]]) -> int:
        row = start_row
        for i, (label, value) in enumerate(data):
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_label.font = LABEL_FONT
            cell_label.border = CELL_BORDER
            cell_label.alignment = Alignment(vertical="top")

            cell_value = ws.cell(row=row, column=2, value=_safe_str(value) if not isinstance(value, (int, float)) else value)
            cell_value.font = DATA_FONT
            cell_value.border = CELL_BORDER

            if isinstance(value, float):
                cell_value.number_format = '#,##0.00'
            elif isinstance(value, int):
                cell_value.number_format = '#,##0'

            if i % 2 == 1:
                cell_label.fill = ALT_ROW_FILL
                cell_value.fill = ALT_ROW_FILL

            row += 1
        return row + 1

    def setup_sheet(ws, col_widths: list[int] | None = None):
        ws.sheet_properties.tabColor = "3B82F6"
        if col_widths:
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()
    content = dossier_data.get("content", {})

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    setup_sheet(ws, [40, 60])

    row = 1
    ws.cell(row=row, column=1, value=dossier_data.get("title", "Dossier Report")).font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Entity: {dossier_data.get('entity', '')}").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Confidence: {dossier_data.get('confidence_score', 0) * 100:.0f}%").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Agents: {dossier_data.get('agents_completed', 0)}/{dossier_data.get('agents_total', 0)}").font = SUBTITLE_FONT
    row += 2

    es = content.get("executive_summary", {})
    if es:
        row = write_header(ws, row, "Executive Summary")
        row = write_kv_rows(ws, row, _flatten_dict(es))

    thesis = content.get("investment_thesis", {})
    if thesis:
        row = write_header(ws, row, "Investment Thesis")
        row = write_kv_rows(ws, row, _flatten_dict(thesis))

    cost = dossier_data.get("cost", {})
    if cost:
        row = write_header(ws, row, "Pipeline Cost")
        row = write_kv_rows(ws, row, [
            ("Total Tokens", cost.get("total_tokens", 0)),
            ("Total Cost USD", cost.get("total_cost_usd", 0.0)),
        ])

    # ── Sheet 2: Financial Statements ──
    ws2 = wb.create_sheet("Financial Statements")
    setup_sheet(ws2, [40, 60])
    row = 1
    fundamentals = content.get("fundamentals", {})
    for section_key, section_label in [
        ("profile", "Company Profile"), ("financials", "Income Statement"),
        ("market_data", "Market Data"), ("revenue_model", "Revenue Model"),
        ("profitability", "Profitability"), ("balance_sheet", "Balance Sheet"),
        ("cash_flow", "Cash Flow"), ("growth", "Growth"),
    ]:
        section_data = fundamentals.get(section_key, {})
        if section_data and isinstance(section_data, dict):
            row = write_header(ws2, row, section_label)
            row = write_kv_rows(ws2, row, _flatten_dict(section_data))

    # ── Sheet 3: Valuation ──
    ws3 = wb.create_sheet("Valuation")
    setup_sheet(ws3, [40, 60])
    row = 1
    valuation = content.get("valuation", {})
    for method_key, method_label in [
        ("dcf", "Discounted Cash Flow (DCF)"), ("comps", "Comparable Companies"),
        ("precedent_transactions", "Precedent Transactions"), ("sum_of_parts", "Sum of Parts"),
    ]:
        method_data = valuation.get(method_key, {})
        if method_data and isinstance(method_data, dict):
            row = write_header(ws3, row, method_label)
            row = write_kv_rows(ws3, row, _flatten_dict(method_data))

    # ── Sheet 4: Risk Assessment ──
    ws4 = wb.create_sheet("Risk Assessment")
    setup_sheet(ws4, [40, 60])
    row = 1
    risk_quality = content.get("risk_and_quality", {})
    for section_key, section_label in [
        ("risk_assessment", "Risk Assessment"), ("competitive_moat", "Competitive Moat"),
        ("esg", "ESG & Governance"), ("management", "Management Quality"),
    ]:
        section_data = risk_quality.get(section_key, {})
        if section_data and isinstance(section_data, dict):
            row = write_header(ws4, row, section_label)
            row = write_kv_rows(ws4, row, _flatten_dict(section_data))

    # ── Sheet 5: All Agent Outputs (raw) ──
    ws5 = wb.create_sheet("All Agent Outputs")
    setup_sheet(ws5, [30, 70])
    row = 1
    ws5.cell(row=row, column=1, value="Section").font = HEADER_FONT
    ws5.cell(row=row, column=2, value="Data").font = HEADER_FONT
    row += 1
    for section_name, section_data in content.items():
        if not section_data or not isinstance(section_data, dict):
            continue
        row = write_header(ws5, row, section_name.replace("_", " ").title())
        row = write_kv_rows(ws5, row, _flatten_dict(section_data))

    for extra_key in ["industry_context", "news_sentiment"]:
        extra_data = content.get(extra_key, {})
        if extra_data and isinstance(extra_data, dict):
            row = write_header(ws5, row, extra_key.replace("_", " ").title())
            row = write_kv_rows(ws5, row, _flatten_dict(extra_data))

    # ── Serialize ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
